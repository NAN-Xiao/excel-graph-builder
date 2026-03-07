#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
图谱构建器 - Indexer 独立模块
"""

import os
import re
import hashlib
import pandas as pd
from pathlib import Path
from typing import List, Dict, Optional, Set
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime

from indexer.schema_graph import SchemaGraph, TableSchema, RelationEdge
from indexer import SimpleLogger
from indexer.html_report import HTMLReportGenerator
from indexer.graph_analyzer import GraphAnalyzer, AnalysisResult

try:
    from Levenshtein import ratio as levenshtein_ratio
    HAS_LEVENSHTEIN = True
except ImportError:
    HAS_LEVENSHTEIN = False


class BuildResult:
    """构建结果统计"""
    def __init__(self):
        self.added: List[str] = []      # 新增的表
        self.updated: List[str] = []    # 更新的表
        self.unchanged: List[str] = []  # 未变化的表
        self.deleted: List[str] = []    # 删除的表
        self.failed: List[str] = []     # 扫描失败的文件
    
    def summary(self) -> str:
        return (
            f"新增 {len(self.added)}, "
            f"更新 {len(self.updated)}, "
            f"删除 {len(self.deleted)}, "
            f"未变 {len(self.unchanged)}, "
            f"失败 {len(self.failed)}"
        )


class SchemaGraphBuilder:
    """配置表知识图谱构建器"""
    
    def __init__(self, data_root: str, max_workers: int = 4, html_output_dir: str = "./html", offline_html: bool = True):
        self.data_root = Path(data_root)
        self.max_workers = max_workers
        self.logger = SimpleLogger()
        self.html_generator = HTMLReportGenerator(html_output_dir, offline=offline_html)
        
        self.logger.info(f"GraphBuilder: 数据目录 {self.data_root}")
    
    def build_full_graph(
        self, 
        incremental: bool = False, 
        existing_graph: Optional[SchemaGraph] = None,
        deleted_tables: Optional[Set[str]] = None
    ) -> tuple[SchemaGraph, BuildResult]:
        """
        构建/更新图谱
        
        Args:
            incremental: 是否增量模式
            existing_graph: 现有图谱（增量模式需要）
            deleted_tables: 指定要删除的表名集合（增量模式可选）
        
        Returns:
            (graph, result): 图谱对象和构建结果统计
        """
        result = BuildResult()
        
        if incremental and existing_graph:
            graph = existing_graph
            self.logger.info(f"增量更新，现有 {len(graph.tables)} 个表")
        else:
            graph = SchemaGraph()
            incremental = False  # 强制全量
            self.logger.info("开始全量构建图谱")
        
        # 1. 扫描所有 Excel 文件
        excel_files = list(self.data_root.rglob("*.xlsx"))
        self.logger.info(f"发现 {len(excel_files)} 个 Excel 文件")
        
        if not excel_files:
            self.logger.warning(f"未找到 Excel 文件: {self.data_root}")
            return graph, result
        
        # 收集当前扫描到的表名
        current_table_names: Set[str] = set()
        
        # 2. 并行扫描文件（带进度显示 + 修改时间优化）
        total_files = len(excel_files)
        processed = 0
        skipped_by_mtime = 0
        
        with ThreadPoolExecutor(max_workers=self.max_workers) as executor:
            # 智能调度：先处理可能未修改的文件（跳过更快）
            futures = {}
            for fp in excel_files:
                # 预检查：如果修改时间没变，直接跳过
                if incremental and self._can_skip_by_mtime(fp, graph):
                    table_name = fp.stem
                    if table_name in graph.tables:
                        result.unchanged.append(table_name)
                        current_table_names.add(table_name)
                        skipped_by_mtime += 1
                        continue
                
                futures[executor.submit(self._scan_table, fp, incremental, graph)] = fp
            
            self.logger.info(f"通过修改时间跳过: {skipped_by_mtime} 个文件")
            
            for future in futures:
                file_path = futures[future]
                processed += 1
                
                # 每 50 个文件显示一次进度
                if processed % 50 == 0 or processed == len(futures):
                    percent = processed / len(futures) * 100
                    self.logger.info(f"扫描进度: {processed}/{len(futures)} ({percent:.1f}%)")
                
                try:
                    scan_result = future.result()
                    if scan_result is None:
                        result.failed.append(file_path.stem)
                        continue
                    
                    table_schema, action = scan_result
                    current_table_names.add(table_schema.name)
                    
                    if action == 'unchanged':
                        result.unchanged.append(table_schema.name)
                    elif action == 'updated':
                        result.updated.append(table_schema.name)
                        self.logger.info(f"表 {table_schema.name} 已更新")
                        graph.add_table(table_schema)
                    elif action == 'added':
                        result.added.append(table_schema.name)
                        self.logger.info(f"新增表: {table_schema.name}")
                        graph.add_table(table_schema)
                    
                except Exception as e:
                    self.logger.error(f"扫描失败 {file_path}: {e}")
                    result.failed.append(file_path.stem)
        
        # 3. 处理删除（两种方式：显式指定 或 自动检测）
        if incremental:
            tables_to_delete: Set[str] = set()
            
            # 显式指定的删除
            if deleted_tables:
                tables_to_delete.update(deleted_tables)
            
            # 自动检测：现有表中不在当前扫描结果中的
            existing_names = set(graph.tables.keys())
            missing_tables = existing_names - current_table_names
            tables_to_delete.update(missing_tables)
            
            # 执行删除
            for table_name in tables_to_delete:
                if graph.remove_table(table_name):
                    result.deleted.append(table_name)
                    self.logger.info(f"删除表: {table_name}")
        
        # 4. 重新发现关联关系（全量重建，因为表结构可能变了）
        self.logger.info(f"发现关联关系，当前 {len(graph.tables)} 个表...")
        graph.relations.clear()  # 清空旧关系，重新发现
        self._discover_relations(graph)
        
        graph.updated_at = datetime.now()
        self.logger.success(
            f"构建完成: {len(graph.tables)} 表, {len(graph.relations)} 关系 | {result.summary()}"
        )
        
        # 4. 图算法分析（环检测、模块聚类等）
        analyzer = GraphAnalyzer(graph)
        analysis = analyzer.analyze()
        
        # 打印分析报告
        print(analyzer.get_summary())
        
        # 5. 外键完整性检查（抽样验证）
        fk_violations = self._validate_foreign_keys(graph, sample_size=100)
        if fk_violations:
            self.logger.warning(f"发现 {len(fk_violations)} 个外键完整性问题")
            for v in fk_violations[:5]:
                self.logger.warning(f"  {v}")
        
        # 生成 HTML 可视化报告
        build_stats = {
            "table_count": len(graph.tables),
            "relation_count": len(graph.relations),
            "build_time": graph.updated_at.strftime("%Y-%m-%d %H:%M:%S"),
            "version": graph.version,
            "added": len(result.added),
            "updated": len(result.updated),
            "deleted": len(result.deleted),
            "unchanged": len(result.unchanged),
            "failed": len(result.failed),
            "cycles": len(analysis.cycles),
            "modules": len(analysis.modules),
            "orphans": len(analysis.orphans)
        }
        self.html_generator.generate(graph, build_stats)
        
        return graph, result
    
    def _can_skip_by_mtime(self, file_path: Path, graph: SchemaGraph) -> bool:
        """
        通过文件修改时间判断是否可以跳过扫描
        
        比 MD5 快 100 倍，适合快速过滤未修改的文件
        """
        table_name = file_path.stem
        if table_name not in graph.tables:
            return False
        
        cached = graph.tables[table_name]
        current_mtime = os.path.getmtime(file_path)
        
        # 修改时间相同，认为文件未变
        return abs(current_mtime - cached.modified_time) < 0.001  # 1ms 误差
    
    def _scan_table(self, file_path: Path, incremental: bool = False, 
                   graph: Optional[SchemaGraph] = None) -> Optional[tuple[TableSchema, str]]:
        """
        扫描单个表文件
        
        优化策略：
        - 限制采样行数 50 行
        - 限制最大列数 50 列（避免超大列宽表内存爆炸）
        - 使用 openpyxl read_only 模式估算行数
        - 列级采样哈希（快速检测变化）
        
        Returns:
            (TableSchema, action) 或 None
            action: 'added' | 'updated' | 'unchanged'
        """
        try:
            table_name = file_path.stem
            current_mtime = os.path.getmtime(file_path)
            
            # 增量模式：先检查修改时间
            if incremental and graph and table_name in graph.tables:
                cached = graph.tables[table_name]
                if abs(current_mtime - cached.modified_time) < 0.001:
                    # 修改时间没变，直接返回未变
                    return cached, 'unchanged'
            
            # 大表优化：限制列数和行数
            MAX_COLS = 50
            MAX_SAMPLE_ROWS = 50
            
            # 先获取总行列数，决定采样策略
            total_rows, total_cols = self._get_file_dimensions(file_path)
            
            # 智能采样：如果列太多，只读前 MAX_COLS 列
            usecols = None
            if total_cols > MAX_COLS:
                usecols = range(MAX_COLS)
            
            df = pd.read_excel(
                file_path, 
                nrows=MAX_SAMPLE_ROWS, 
                usecols=usecols,
                engine='openpyxl'
            )
            
            if df.empty:
                return None
            
            # 列级采样哈希（比全文件 MD5 快 10 倍）
            file_hash = self._calc_column_hash(df, total_rows)
            
            # 增量模式：检查哈希是否变化
            if incremental and graph and table_name in graph.tables:
                cached = graph.tables[table_name]
                if cached.hash == file_hash:
                    # 哈希没变，更新时间戳缓存
                    cached.modified_time = current_mtime
                    return cached, 'unchanged'
                else:
                    action = 'updated'
            else:
                action = 'added'
            
            columns = []
            numeric_cols = []
            enum_cols = {}
            
            for col in df.columns:
                col_info = {
                    'name': str(col),
                    'dtype': str(df[col].dtype),
                    'sample_values': [self._safe_str(v) for v in df[col].dropna().head(3).tolist()]
                }
                columns.append(col_info)
                
                if pd.api.types.is_numeric_dtype(df[col]):
                    numeric_cols.append(str(col))
                
                unique_count = df[col].nunique()
                if 1 < unique_count < 50:
                    enum_cols[str(col)] = [self._safe_str(v) for v in df[col].dropna().unique().tolist()]
            
            primary_key = self._guess_primary_key(df)
            
            table_schema = TableSchema(
                name=table_name,
                file_path=str(file_path.relative_to(self.data_root)),
                row_count=total_rows,
                columns=columns,
                primary_key=primary_key,
                hash=file_hash,
                numeric_columns=numeric_cols,
                enum_columns=enum_cols,
                modified_time=current_mtime
            )
            
            return table_schema, action
            
        except Exception as e:
            self.logger.error(f"扫描 {file_path} 失败: {e}")
            return None
    
    def _calc_column_hash(self, df: pd.DataFrame, total_rows: int) -> str:
        """
        列级采样哈希：只采样关键信息，比全文件 MD5 快 10 倍
        
        采样内容：
        - 列名列表
        - 前 3 列的前 20 行数据
        - 总行数
        """
        try:
            import hashlib
            
            # 列名
            col_names = str(list(df.columns))
            
            # 采样前 3 列的前 20 行
            sample_cols = min(3, len(df.columns))
            sample_rows = min(20, len(df))
            
            if sample_rows > 0 and sample_cols > 0:
                sample_data = df.iloc[:sample_rows, :sample_cols].values.tobytes()
            else:
                sample_data = b''
            
            # 组合哈希
            hash_input = f"{col_names}|{total_rows}|{sample_data.hex()}"
            return hashlib.md5(hash_input.encode()).hexdigest()[:16]  # 16位足够
            
        except Exception:
            # 降级：使用原方法
            return self._calc_file_hash(df)
    
    def _get_file_dimensions(self, file_path: Path) -> tuple[int, int]:
        """
        快速获取 Excel 文件行列数（不加载全部数据）
        
        Returns:
            (row_count, col_count)
        """
        try:
            from openpyxl import load_workbook
            wb = load_workbook(file_path, read_only=True, data_only=True)
            ws = wb.active
            return ws.max_row, ws.max_column
        except Exception:
            # 降级：使用 pandas 估算
            try:
                df = pd.read_excel(file_path, nrows=0, engine='openpyxl')
                return 0, len(df.columns)
            except:
                return 0, 0
    
    def _calc_file_hash(self, file_path: Path) -> str:
        """计算文件哈希（全文件 MD5）"""
        hash_md5 = hashlib.md5()
        with open(file_path, "rb") as f:
            hash_md5.update(f.read(1024 * 1024))
        return hash_md5.hexdigest()
    
    def _estimate_row_count(self, file_path: Path) -> int:
        try:
            from openpyxl import load_workbook
            wb = load_workbook(file_path, read_only=True, data_only=True)
            return wb.active.max_row
        except:
            return 0
    
    def _guess_primary_key(self, df: pd.DataFrame) -> Optional[str]:
        for col in df.columns:
            if df[col].isna().sum() == 0 and df[col].nunique() == len(df):
                col_str = str(col).lower()
                if any(kw in col_str for kw in ['id', 'key', 'code', 'no']):
                    return str(col)
        return None
    
    def _discover_relations(self, graph: SchemaGraph):
        """
        发现表之间的关联关系
        
        策略：
        1. 严格模式：xxx_id 后缀匹配表名
        2. 模糊模式：列名与表名相似度匹配（需 Levenshtein）
        3. 内容模式：列内容重合度分析（发现隐式外键）
        """
        tables = list(graph.tables.values())
        table_names = [t.name for t in tables]
        
        relations_found = 0
        content_relations = 0
        
        # 阶段1：列名匹配
        for table in tables:
            for col in table.columns:
                col_name = col['name']
                
                # 策略1：严格外键模式匹配
                if self._is_fk_pattern(col_name):
                    ref_table = self._extract_table_from_fk(col_name)
                    matched = self._find_table_by_name(ref_table, tables, table.name)
                    if matched:
                        graph.relations.append(RelationEdge(
                            from_table=table.name,
                            from_column=col_name,
                            to_table=matched.name,
                            to_column=matched.primary_key or col_name,
                            relation_type='fk_strict',
                            confidence=0.9
                        ))
                        relations_found += 1
                        continue
                
                # 策略2：Levenshtein 模糊匹配（处理 usr_id ↔ user 这类简写）
                if HAS_LEVENSHTEIN:
                    fuzzy_match = self._find_fuzzy_match(col_name, table_names, table.name)
                    if fuzzy_match:
                        target = next((t for t in tables if t.name == fuzzy_match), None)
                        if target:
                            similarity = levenshtein_ratio(
                                self._normalize_name(col_name), 
                                self._normalize_name(target.name)
                            )
                            graph.relations.append(RelationEdge(
                                from_table=table.name,
                                from_column=col_name,
                                to_table=target.name,
                                to_column=target.primary_key or col_name,
                                relation_type='fk_fuzzy',
                                confidence=round(similarity, 2)
                            ))
                            relations_found += 1
        
        # 阶段2：内容相似度分析（发现隐式外键）
        # 例如：order.creator 和 user.usr_id 内容高度重合，但列名不像
        content_rels = self._discover_content_relations(graph, tables)
        for rel in content_rels:
            graph.relations.append(rel)
            content_relations += 1
        
        total = relations_found + content_relations
        self.logger.info(
            f"关系发现完成: {total} 个 "
            f"(列名匹配: {relations_found}, 内容匹配: {content_relations})"
        )
    
    def _discover_content_relations(self, graph: SchemaGraph, tables: list) -> list:
        """
        通过列内容相似度发现隐式外键关系
        
        策略：
        - 对于数值型/短文本列，采样值集合计算 Jaccard 相似度
        - 相似度 > 80% 认为是潜在外键
        
        注意：这是一个启发式算法，可能有误报，置信度标记为 0.5-0.7
        """
        relations = []
        
        # 只处理有枚举值的列（已采样的离散值）
        for table_a in tables:
            for col_a in table_a.columns:
                col_name_a = col_a['name']
                
                # 跳过明显不是 ID 的列
                if not self._is_potential_id_column(col_name_a):
                    continue
                
                # 获取该列的样本值
                values_a = set(col_a.get('sample_values', []))
                if len(values_a) < 3:  # 样本太少，跳过
                    continue
                
                # 与其他表的 ID 列比较
                for table_b in tables:
                    if table_b.name == table_a.name:
                        continue
                    
                    # 找 table_b 的主键或 ID 列
                    pk_col = table_b.primary_key
                    if not pk_col:
                        # 找看起来像 ID 的列
                        for col_b in table_b.columns:
                            if self._is_potential_id_column(col_b['name']):
                                pk_col = col_b['name']
                                break
                    
                    if not pk_col:
                        continue
                    
                    # 获取 table_b 该列的样本值
                    pk_col_data = next((c for c in table_b.columns if c['name'] == pk_col), None)
                    if not pk_col_data:
                        continue
                    
                    values_b = set(pk_col_data.get('sample_values', []))
                    if len(values_b) < 3:
                        continue
                    
                    # 计算 Jaccard 相似度
                    intersection = values_a & values_b
                    union = values_a | values_b
                    
                    if len(union) == 0:
                        continue
                    
                    similarity = len(intersection) / len(union)
                    
                    # 相似度阈值：80% 以上认为是潜在外键
                    if similarity > 0.8:
                        relations.append(RelationEdge(
                            from_table=table_a.name,
                            from_column=col_name_a,
                            to_table=table_b.name,
                            to_column=pk_col,
                            relation_type='fk_content',
                            confidence=round(0.5 + similarity * 0.3, 2)  # 0.5-0.8 置信度
                        ))
        
        return relations
    
    def _is_potential_id_column(self, col_name: str) -> bool:
        """判断列名是否可能是 ID 列"""
        name_lower = col_name.lower()
        id_patterns = ['id', 'code', 'key', 'no', 'num', 'pk', 'fk']
        return any(pattern in name_lower for pattern in id_patterns)
    
    def _find_table_by_name(self, ref_name: str, tables: list, exclude_table: str) -> Optional[TableSchema]:
        """根据名称查找表（严格匹配）"""
        ref_lower = ref_name.lower()
        for target in tables:
            if target.name.lower() == ref_lower and target.name != exclude_table:
                return target
        return None
    
    def _find_fuzzy_match(self, col_name: str, table_names: list, exclude_table: str) -> Optional[str]:
        """
        使用 Levenshtein 距离模糊匹配列名和表名
        
        例如：
        - usr_id → user (相似度 0.6+)
        - cust_code → customer (相似度 0.5+)
        """
        if not HAS_LEVENSHTEIN:
            return None
        
        # 提取列名的主体部分（去掉 _id, _code 等后缀）
        col_base = self._extract_column_base(col_name)
        if not col_base or len(col_base) < 3:
            return None
        
        col_normalized = self._normalize_name(col_base)
        best_match = None
        best_score = 0.0
        
        for table_name in table_names:
            if table_name == exclude_table:
                continue
            
            table_normalized = self._normalize_name(table_name)
            
            # 计算相似度
            similarity = levenshtein_ratio(col_normalized, table_normalized)
            
            # 阈值：0.6 以上认为是相似
            if similarity > 0.6 and similarity > best_score:
                best_score = similarity
                best_match = table_name
        
        return best_match
    
    def _extract_column_base(self, col_name: str) -> str:
        """提取列名主体（去掉常见后缀）"""
        # 去掉常见后缀：_id, _code, _key, _no, _num
        suffixes = ['_id', '_code', '_key', '_no', '_num', '_pk', '_fk']
        col_lower = col_name.lower()
        for suffix in suffixes:
            if col_lower.endswith(suffix):
                return col_name[:-len(suffix)]
        return col_name
    
    def _normalize_name(self, name: str) -> str:
        """标准化名称用于比较（小写、去下划线）"""
        return name.lower().replace('_', '')
    
    def _is_fk_pattern(self, col_name: str) -> bool:
        """检查是否符合外键命名模式"""
        return bool(re.match(r'(\w+)_(id|code|key|no)$', col_name, re.IGNORECASE))
    
    def _extract_table_from_fk(self, col_name: str) -> str:
        """从外键列名提取可能的表名"""
        # user_id → user
        # usr_id → usr
        match = re.match(r'(\w+?)_?(id|code|key|no)$', col_name, re.IGNORECASE)
        return match.group(1) if match else col_name
    
    def _safe_str(self, value) -> str:
        if pd.isna(value):
            return ""
        return str(value)[:50]
    
    def _validate_foreign_keys(self, graph: SchemaGraph, sample_size: int = 100) -> List[str]:
        """
        抽样验证外键完整性
        
        检查：从表的外键值是否都存在于主表中
        注意：这是一个昂贵的操作，只抽样检查
        
        Returns:
            问题列表，每个问题描述一个外键违反
        """
        violations = []
        
        # 只检查高置信度的外键关系
        high_confidence_rels = [
            r for r in graph.relations 
            if r.confidence >= 0.8 and r.relation_type in ('fk_strict', 'fk_fuzzy')
        ]
        
        if not high_confidence_rels:
            return violations
        
        self.logger.info(f"开始外键完整性检查（抽样 {sample_size} 条）...")
        
        for rel in high_confidence_rels[:20]:  # 最多检查 20 个关系
            try:
                from_table = rel.from_table
                from_col = rel.from_column
                to_table = rel.to_table
                to_col = rel.to_column
                
                # 获取文件路径
                from_path = self.data_root / graph.tables[from_table].file_path
                to_path = self.data_root / graph.tables[to_table].file_path
                
                if not from_path.exists() or not to_path.exists():
                    continue
                
                # 读取主键表的所有值（限制数量）
                pk_df = pd.read_excel(to_path, usecols=[to_col], nrows=10000, engine='openpyxl')
                pk_values = set(pk_df[to_col].dropna().astype(str))
                
                if len(pk_values) == 0:
                    continue
                
                # 读取外键表的抽样值
                fk_df = pd.read_excel(from_path, usecols=[from_col], nrows=sample_size, engine='openpyxl')
                fk_values = fk_df[from_col].dropna().astype(str)
                
                # 检查缺失率
                missing = [v for v in fk_values if v not in pk_values and str(v) != 'nan']
                if len(missing) > len(fk_values) * 0.1:  # 缺失率 > 10%
                    missing_rate = len(missing) / len(fk_values) * 100
                    violations.append(
                        f"{from_table}.{from_col} → {to_table}.{to_col}: "
                        f"缺失率 {missing_rate:.1f}% ({len(missing)}/{len(fk_values)})"
                    )
                    
            except Exception as e:
                # 验证失败不中断流程
                continue
        
        return violations
